from openpyxl import workbook, load_workbook

try:
    wb = load_workbook('database.xlsx')
    ws = wb.active
except FileNotFoundError:
    wb = workbook()
    ws = wb.active
    titles = ["ID", "Name", "Phone", "Address", "ID Proof", "ID Number", "Password"]
    ws.append(titles)

a=True
while a is True:
    print("-----------------------")
    print("------ADMIN USER-------")
    print("-----------------------")
    print("1.Add New Employee")
    print("2.Find the Emplyee by id")
    print("3.Delete employee")
    print("4.Update the information")
    print("5.Log Out")
    print("-----------------------")
    opt=input("enter the option:")
    if opt=='1':
        print("---------------------------------------------------")
        no=int(input("enter the number of employees to add:"))
        current_id = len(list(ws.iter_rows()))
        for i in range(no):
            name=input("enter the name:")
            phn=input("enter the phone number:")
            adrs=input("enter the address:")
            id_pr=input("enter the id proof:")
            id_num=input("enter the id number:")
            passw=input("enter the password:")
            ws.append([str(current_id), name, phn, adrs, id_pr, id_num, passw])
            current_id+=1
        print("---------------------------------------------------")
    wb.save('database.xlsx')
    if opt=='2':
        print("---------------------------------------------------")
        id_find=input("enter the id to find:")
        found=False
        for row_num,i in enumerate(ws.iter_rows(min_row=2, values_only=True),start=2):
            id=i[0]
            if id==id_find:
                print("Employee details are ")
                for i, cell in enumerate(i):
                    titles = ["ID", "Name", "Phone", "Address", "ID Proof", "ID Number", "Password"]
                    print(titles[i] + ": " + str(cell))
                found = True
                break
        if not found:
            print("not found")
        wb.save('database.xlsx')
        print("---------------------------------------------------")
    if opt=='3':
        print("---------------------------------------------------")
        rem=input("enter the id to delete the employee:")
        found=False
        for row_num,i in enumerate(ws.iter_rows(min_row=2, values_only=True),start=2):
            id_rem=i[0]
            if id_rem==rem:
                ws.delete_rows(row_num)
                found=True
                wb.save('database.xlsx')
                print("employee deleted successfully")
                break
        if not found:
            print("not found")   
        print("---------------------------------------------------")
    if opt=='4':
        print("---------------------------------------------------")
        try:
            titles = ["ID", "Name", "Phone", "Address", "ID Proof", "ID Number", "Password"]
            id_up=input("enter the id to update:")
            up=input("enter what to change:").strip().title()
            chng=input("enter the changes:")
            found=False
            for row_num,i in enumerate(ws.iter_rows(min_row=2, values_only=True),start=2):
                id_u=i[0]
                if i[0] == id_up:
                    if up in titles:
                        col_num = titles.index(up) + 1
                        ws.cell(row=row_num, column=col_num, value=chng)
                        wb.save('database.xlsx')
                        print(f"{up} updated successfully.")
                        found = True
                        break
            if not found:
                print("Employee not found")
        except ValueError:
            print("Invalid Input")
        print("---------------------------------------------------")
    if opt=='5':
        a=False
wb.save('database.xlsx')