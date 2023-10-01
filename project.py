from openpyxl import load_workbook
try:
    wb = load_workbook('database.xlsx')
    sheet1 = wb['sheet1']
except FileNotFoundError:
    print("Database file not found.")

def admin_login():
    import admin
    while True:
        opt=input("logout or exit: ")
        if opt=='logout':
            return
        elif opt=='exit':
            exit()
            
def emp_login(username, password):
    found=False
    for row_num,row in enumerate(sheet1.iter_rows(min_row=0, values_only=True),start=2):
        username_in_sheet1, password_in_sheet1 = row[0], row[6]
        if username_in_sheet1 == username and password_in_sheet1 == password:
            found = True
            break
    if found:
        import employee
        while True:
            opt=input("logout or exit: ")
            if opt=='logout':
                return True
            elif opt=='exit':
                exit()
    else:
        print("invalid credentials")
        return False

while True:
    print("-------------------------------------------")
    username=input("enter the username: ")
    password=input("enter the password: ")
    print("-------------------------------------------")
    
    if username == 'admin' and password == 'admin':
        print("-------------------------------------------")
        print("Login Successful")
        print("-------------------------------------------")
        admin_login()

    else:
        print("-------------------------------------------")
        print("Login Successful")
        print("-------------------------------------------")
        emp_login(username, password)
            

