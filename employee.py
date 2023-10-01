from openpyxl import workbook, load_workbook
from datetime import datetime

db_wb = load_workbook('database.xlsx')
db_sh = db_wb['sheet1']
att_wb = load_workbook('Attendence.xlsx')
att_ws = att_wb.active
print("---------------------------------------------------")
emp_id=input("enter the employee id:")
id_found = False
for row in db_sh.iter_rows(min_row=2, values_only=True):
    id_in_sheet1 = row[0]
    if id_in_sheet1 == emp_id:
        id_found = True
        employee_name = row[1]
        break
print("---------------------------------------------------")
if id_found:
    atdn = input("Enter the attendance: ")
    
    if atdn == 'present':
        current_date = datetime.now().strftime('%Y-%m-%d')
        att_ws.append([emp_id, employee_name, 'present', current_date])
        att_wb.save('Attendence.xlsx')
        print("Attendance recorded successfully.")
    else:
        print("Attendance not recorded.")
else:
    print("Employee ID not found in sheet1. Please request admin sign-in.")
print("---------------------------------------------------")